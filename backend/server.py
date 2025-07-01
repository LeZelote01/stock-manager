from fastapi import FastAPI, APIRouter, HTTPException, status, BackgroundTasks, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
import bcrypt
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
import joblib
import json
import io
import base64
import qrcode
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
import plotly.express as px
import socketio
import asyncio
import aiofiles
import csv
from io import StringIO, BytesIO
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="Advanced Stock Manager API", version="2.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Socket.IO for real-time notifications
sio = socketio.AsyncServer(cors_allowed_origins="*", async_mode='asgi')
socket_app = socketio.ASGIApp(sio)

# Models - Enhanced versions of existing models
class Material(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    quantite: int
    date_ajout: datetime = Field(default_factory=datetime.utcnow)
    qr_code: Optional[str] = None
    prix_unitaire: Optional[float] = 0.0
    seuil_alerte: Optional[int] = 10
    seuil_critique: Optional[int] = 5
    fournisseur_id: Optional[str] = None
    categorie: Optional[str] = "Général"
    unite: Optional[str] = "unité"
    emplacement: Optional[str] = ""
    
class MaterialCreate(BaseModel):
    nom: str
    quantite: int
    prix_unitaire: Optional[float] = 0.0
    seuil_alerte: Optional[int] = 10
    seuil_critique: Optional[int] = 5
    fournisseur_id: Optional[str] = None
    categorie: Optional[str] = "Général"
    unite: Optional[str] = "unité"
    emplacement: Optional[str] = ""

class MaterialUpdate(BaseModel):
    nom: Optional[str] = None
    quantite: Optional[int] = None
    prix_unitaire: Optional[float] = None
    seuil_alerte: Optional[int] = None
    seuil_critique: Optional[int] = None
    fournisseur_id: Optional[str] = None
    categorie: Optional[str] = None
    unite: Optional[str] = None
    emplacement: Optional[str] = None

class Agent(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class AgentCreate(BaseModel):
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class Superviseur(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class SuperviseurCreate(BaseModel):
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class ChefSection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class ChefSectionCreate(BaseModel):
    nom: str
    matricule: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    service: Optional[str] = ""

class DemandeSortie(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    superviseur_id: str
    superviseur_nom: str
    superviseur_matricule: str
    agent1_id: str
    agent1_nom: str
    agent1_matricule: str
    agent2_id: str
    agent2_nom: str
    agent2_matricule: str
    date: datetime = Field(default_factory=datetime.utcnow)
    materiels_demandes: Dict[str, int]  # material_id -> quantite
    signature: Optional[str] = None
    status: str = "approuve"  # Auto-approved as requested
    notes: Optional[str] = ""
    valeur_totale: Optional[float] = 0.0

class DemandeSortieCreate(BaseModel):
    superviseur_id: str
    agent1_id: str
    agent2_id: str
    materiels_demandes: Dict[str, int]
    signature: Optional[str] = None
    notes: Optional[str] = ""

class Fournisseur(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    contact: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    adresse: Optional[str] = ""
    delai_livraison: Optional[int] = 7  # en jours

class FournisseurCreate(BaseModel):
    nom: str
    contact: str
    email: Optional[str] = None
    telephone: Optional[str] = None
    adresse: Optional[str] = ""
    delai_livraison: Optional[int] = 7

class HistoriqueAction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: datetime = Field(default_factory=datetime.utcnow)
    utilisateur: str
    action: str
    details: Dict[str, Any]
    table_affectee: str

class AdminAuth(BaseModel):
    password: str

class ExportRequest(BaseModel):
    format: str  # pdf, excel, csv, png
    date_debut: Optional[datetime] = None
    date_fin: Optional[datetime] = None
    include_graphiques: Optional[bool] = True

# Admin password (in production, use proper authentication)
ADMIN_PASSWORD = "admin123"

# AI/ML Components for stock prediction
class StockPredictor:
    def __init__(self):
        self.model = RandomForestRegressor(n_estimators=100, random_state=42)
        self.scaler = StandardScaler()
        self.trained = False
    
    async def train_model(self):
        """Train the prediction model with historical data"""
        try:
            # Get historical demand data
            demandes = await db.demandes_sortie.find().to_list(1000)
            
            if len(demandes) < 10:
                return False
            
            # Prepare training data
            df_data = []
            for demande in demandes:
                for material_id, quantite in demande['materiels_demandes'].items():
                    df_data.append({
                        'date': demande['date'],
                        'material_id': material_id,
                        'quantite': quantite,
                        'month': demande['date'].month,
                        'day_of_week': demande['date'].weekday(),
                        'week_of_year': demande['date'].isocalendar()[1]
                    })
            
            if not df_data:
                return False
                
            df = pd.DataFrame(df_data)
            
            # Create features for each material
            features = []
            targets = []
            
            for material_id in df['material_id'].unique():
                material_data = df[df['material_id'] == material_id].sort_values('date')
                
                if len(material_data) < 3:
                    continue
                    
                # Create lag features
                material_data['lag_1'] = material_data['quantite'].shift(1)
                material_data['rolling_mean'] = material_data['quantite'].rolling(3, min_periods=1).mean()
                
                # Remove rows with NaN
                material_data = material_data.dropna()
                
                if len(material_data) < 2:
                    continue
                
                for _, row in material_data.iterrows():
                    features.append([
                        row['month'],
                        row['day_of_week'], 
                        row['week_of_year'],
                        row['lag_1'],
                        row['rolling_mean']
                    ])
                    targets.append(row['quantite'])
            
            if len(features) < 5:
                return False
                
            X = np.array(features)
            y = np.array(targets)
            
            # Scale features
            X_scaled = self.scaler.fit_transform(X)
            
            # Train model
            self.model.fit(X_scaled, y)
            self.trained = True
            
            return True
            
        except Exception as e:
            logging.error(f"Error training model: {e}")
            return False
    
    async def predict_demand(self, material_id: str, days_ahead: int = 30):
        """Predict demand for a material"""
        if not self.trained:
            await self.train_model()
        
        if not self.trained:
            return {"error": "Insufficient data for prediction"}
        
        try:
            # Get recent data for this material
            demandes = await db.demandes_sortie.find({
                f"materiels_demandes.{material_id}": {"$exists": True}
            }).sort("date", -1).limit(5).to_list(5)
            
            if not demandes:
                return {"prediction": 0, "confidence": "low"}
            
            # Calculate recent average
            recent_quantities = []
            for demande in demandes:
                if material_id in demande['materiels_demandes']:
                    recent_quantities.append(demande['materiels_demandes'][material_id])
            
            if not recent_quantities:
                return {"prediction": 0, "confidence": "low"}
            
            # Create features for prediction
            now = datetime.utcnow()
            lag_1 = recent_quantities[0] if recent_quantities else 0
            rolling_mean = np.mean(recent_quantities)
            
            features = np.array([[
                now.month,
                now.weekday(),
                now.isocalendar()[1],
                lag_1,
                rolling_mean
            ]])
            
            # Scale and predict
            features_scaled = self.scaler.transform(features)
            prediction = self.model.predict(features_scaled)[0]
            
            # Adjust for days ahead
            daily_avg = prediction / 30
            final_prediction = daily_avg * days_ahead
            
            return {
                "prediction": max(0, final_prediction),
                "confidence": "medium" if len(recent_quantities) >= 3 else "low"
            }
            
        except Exception as e:
            logging.error(f"Error predicting demand: {e}")
            return {"error": "Prediction failed"}

# Initialize predictor
stock_predictor = StockPredictor()

# Utility functions
async def log_action(utilisateur: str, action: str, details: Dict[str, Any], table_affectee: str):
    """Log user actions for audit trail"""
    historique = HistoriqueAction(
        utilisateur=utilisateur,
        action=action,
        details=details,
        table_affectee=table_affectee
    )
    await db.historique_actions.insert_one(historique.dict())

async def send_notification(message: str, type: str = "info", data: Dict = None):
    """Send real-time notification via WebSocket"""
    notification = {
        "message": message,
        "type": type,
        "timestamp": datetime.utcnow().isoformat(),
        "data": data or {}
    }
    await sio.emit("notification", notification)

def generate_qr_code(data: str) -> str:
    """Generate QR code and return as base64 string"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

# Routes

# WebSocket events
@sio.event
async def connect(sid, environ):
    print(f"Client {sid} connected")
    await sio.emit("connected", {"status": "connected"}, room=sid)

@sio.event
async def disconnect(sid):
    print(f"Client {sid} disconnected")

# Authentication
@api_router.post("/login")
async def login(auth: AdminAuth):
    if auth.password == ADMIN_PASSWORD:
        await log_action("admin", "login", {"success": True}, "auth")
        return {"status": "success", "role": "admin", "token": "admin-token"}
    else:
        await log_action("unknown", "login_failed", {"password_attempt": True}, "auth")
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

# Enhanced Materials CRUD with QR codes
@api_router.get("/materials", response_model=List[Material])
async def get_materials():
    materials = await db.materials.find().to_list(1000)
    return [Material(**material) for material in materials]

@api_router.post("/materials", response_model=Material)
async def create_material(material: MaterialCreate):
    material_dict = material.dict()
    material_obj = Material(**material_dict)
    
    # Generate QR code
    qr_data = f"MAT-{material_obj.id}"
    material_obj.qr_code = generate_qr_code(qr_data)
    
    await db.materials.insert_one(material_obj.dict())
    await log_action("admin", "create_material", {"material_id": material_obj.id, "nom": material_obj.nom}, "materials")
    await send_notification(f"Nouveau matériel ajouté: {material_obj.nom}", "success")
    
    return material_obj

@api_router.put("/materials/{material_id}", response_model=Material)
async def update_material(material_id: str, material_update: MaterialUpdate):
    update_data = {k: v for k, v in material_update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    result = await db.materials.update_one(
        {"id": material_id}, 
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Matériel non trouvé")
    
    updated_material = await db.materials.find_one({"id": material_id})
    await log_action("admin", "update_material", {"material_id": material_id, "updates": update_data}, "materials")
    
    return Material(**updated_material)

@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str):
    # Get material info before deletion
    material = await db.materials.find_one({"id": material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Matériel non trouvé")
    
    result = await db.materials.delete_one({"id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Matériel non trouvé")
    
    await log_action("admin", "delete_material", {"material_id": material_id, "nom": material["nom"]}, "materials")
    await send_notification(f"Matériel supprimé: {material['nom']}", "warning")
    
    return {"message": "Matériel supprimé avec succès"}

# Enhanced Agents CRUD
@api_router.get("/agents", response_model=List[Agent])
async def get_agents():
    agents = await db.agents.find().to_list(1000)
    return [Agent(**agent) for agent in agents]

@api_router.post("/agents", response_model=Agent)
async def create_agent(agent: AgentCreate):
    agent_dict = agent.dict()
    agent_obj = Agent(**agent_dict)
    await db.agents.insert_one(agent_obj.dict())
    await log_action("admin", "create_agent", {"agent_id": agent_obj.id, "nom": agent_obj.nom}, "agents")
    return agent_obj

@api_router.put("/agents/{agent_id}", response_model=Agent)
async def update_agent(agent_id: str, agent_update: AgentCreate):
    result = await db.agents.update_one(
        {"id": agent_id}, 
        {"$set": agent_update.dict()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agent non trouvé")
    
    updated_agent = await db.agents.find_one({"id": agent_id})
    await log_action("admin", "update_agent", {"agent_id": agent_id}, "agents")
    return Agent(**updated_agent)

@api_router.delete("/agents/{agent_id}")
async def delete_agent(agent_id: str):
    agent = await db.agents.find_one({"id": agent_id})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent non trouvé")
    
    result = await db.agents.delete_one({"id": agent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agent non trouvé")
    
    await log_action("admin", "delete_agent", {"agent_id": agent_id, "nom": agent["nom"]}, "agents")
    return {"message": "Agent supprimé avec succès"}

# Enhanced Superviseurs CRUD
@api_router.get("/superviseurs", response_model=List[Superviseur])
async def get_superviseurs():
    superviseurs = await db.superviseurs.find().to_list(1000)
    return [Superviseur(**superviseur) for superviseur in superviseurs]

@api_router.post("/superviseurs", response_model=Superviseur)
async def create_superviseur(superviseur: SuperviseurCreate):
    superviseur_dict = superviseur.dict()
    superviseur_obj = Superviseur(**superviseur_dict)
    await db.superviseurs.insert_one(superviseur_obj.dict())
    await log_action("admin", "create_superviseur", {"superviseur_id": superviseur_obj.id, "nom": superviseur_obj.nom}, "superviseurs")
    return superviseur_obj

@api_router.put("/superviseurs/{superviseur_id}", response_model=Superviseur)
async def update_superviseur(superviseur_id: str, superviseur_update: SuperviseurCreate):
    result = await db.superviseurs.update_one(
        {"id": superviseur_id}, 
        {"$set": superviseur_update.dict()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Superviseur non trouvé")
    
    updated_superviseur = await db.superviseurs.find_one({"id": superviseur_id})
    await log_action("admin", "update_superviseur", {"superviseur_id": superviseur_id}, "superviseurs")
    return Superviseur(**updated_superviseur)

@api_router.delete("/superviseurs/{superviseur_id}")
async def delete_superviseur(superviseur_id: str):
    superviseur = await db.superviseurs.find_one({"id": superviseur_id})
    if not superviseur:
        raise HTTPException(status_code=404, detail="Superviseur non trouvé")
    
    result = await db.superviseurs.delete_one({"id": superviseur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Superviseur non trouvé")
    
    await log_action("admin", "delete_superviseur", {"superviseur_id": superviseur_id, "nom": superviseur["nom"]}, "superviseurs")
    return {"message": "Superviseur supprimé avec succès"}

# Enhanced Chef Section CRUD
@api_router.get("/chef-section", response_model=List[ChefSection])
async def get_chef_section():
    chefs = await db.chef_section.find().to_list(1000)
    return [ChefSection(**chef) for chef in chefs]

@api_router.post("/chef-section", response_model=ChefSection)
async def create_chef_section(chef: ChefSectionCreate):
    chef_dict = chef.dict()
    chef_obj = ChefSection(**chef_dict)
    await db.chef_section.insert_one(chef_obj.dict())
    await log_action("admin", "create_chef_section", {"chef_id": chef_obj.id, "nom": chef_obj.nom}, "chef_section")
    return chef_obj

@api_router.put("/chef-section/{chef_id}", response_model=ChefSection)
async def update_chef_section(chef_id: str, chef_update: ChefSectionCreate):
    result = await db.chef_section.update_one(
        {"id": chef_id}, 
        {"$set": chef_update.dict()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Chef de section non trouvé")
    
    updated_chef = await db.chef_section.find_one({"id": chef_id})
    await log_action("admin", "update_chef_section", {"chef_id": chef_id}, "chef_section")
    return ChefSection(**updated_chef)

@api_router.delete("/chef-section/{chef_id}")
async def delete_chef_section(chef_id: str):
    chef = await db.chef_section.find_one({"id": chef_id})
    if not chef:
        raise HTTPException(status_code=404, detail="Chef de section non trouvé")
    
    result = await db.chef_section.delete_one({"id": chef_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Chef de section non trouvé")
    
    await log_action("admin", "delete_chef_section", {"chef_id": chef_id, "nom": chef["nom"]}, "chef_section")
    return {"message": "Chef de section supprimé avec succès"}

# Enhanced Demandes de sortie with auto-approval
@api_router.get("/demandes", response_model=List[DemandeSortie])
async def get_demandes():
    demandes = await db.demandes_sortie.find().sort("date", -1).to_list(1000)
    return [DemandeSortie(**demande) for demande in demandes]

@api_router.post("/demandes", response_model=DemandeSortie)
async def create_demande(demande_create: DemandeSortieCreate):
    # Get supervisor info
    superviseur = await db.superviseurs.find_one({"id": demande_create.superviseur_id})
    if not superviseur:
        raise HTTPException(status_code=404, detail="Superviseur non trouvé")
    
    # Get agent1 info
    agent1 = await db.agents.find_one({"id": demande_create.agent1_id})
    if not agent1:
        raise HTTPException(status_code=404, detail="Agent 1 non trouvé")
    
    # Get agent2 info
    agent2 = await db.agents.find_one({"id": demande_create.agent2_id})
    if not agent2:
        raise HTTPException(status_code=404, detail="Agent 2 non trouvé")
    
    # Calculate total value
    valeur_totale = 0
    for material_id, quantite in demande_create.materiels_demandes.items():
        material = await db.materials.find_one({"id": material_id})
        if material and quantite > 0:
            valeur_totale += material.get("prix_unitaire", 0) * quantite
    
    # Create demande with full info
    demande_dict = demande_create.dict()
    demande_dict.update({
        "superviseur_nom": superviseur["nom"],
        "superviseur_matricule": superviseur["matricule"],
        "agent1_nom": agent1["nom"],
        "agent1_matricule": agent1["matricule"],
        "agent2_nom": agent2["nom"],
        "agent2_matricule": agent2["matricule"],
        "valeur_totale": valeur_totale,
        "status": "approuve"  # Auto-approved as requested
    })
    
    demande_obj = DemandeSortie(**demande_dict)
    await db.demandes_sortie.insert_one(demande_obj.dict())
    
    # Update stock quantities
    stock_updates = []
    for material_id, quantite in demande_create.materiels_demandes.items():
        if quantite > 0:
            result = await db.materials.update_one(
                {"id": material_id},
                {"$inc": {"quantite": -quantite}}
            )
            stock_updates.append({"material_id": material_id, "quantite_sortie": quantite})
    
    await log_action("user", "create_demande", {
        "demande_id": demande_obj.id,
        "superviseur": superviseur["nom"],
        "valeur_totale": valeur_totale,
        "stock_updates": stock_updates
    }, "demandes_sortie")
    
    await send_notification(f"Nouvelle demande approuvée automatiquement - Valeur: {valeur_totale:.2f}€", "success", {
        "demande_id": demande_obj.id,
        "superviseur": superviseur["nom"]
    })
    
    return demande_obj

# Fournisseurs CRUD
@api_router.get("/fournisseurs", response_model=List[Fournisseur])
async def get_fournisseurs():
    fournisseurs = await db.fournisseurs.find().to_list(1000)
    return [Fournisseur(**fournisseur) for fournisseur in fournisseurs]

@api_router.post("/fournisseurs", response_model=Fournisseur)
async def create_fournisseur(fournisseur: FournisseurCreate):
    fournisseur_dict = fournisseur.dict()
    fournisseur_obj = Fournisseur(**fournisseur_dict)
    await db.fournisseurs.insert_one(fournisseur_obj.dict())
    await log_action("admin", "create_fournisseur", {"fournisseur_id": fournisseur_obj.id, "nom": fournisseur_obj.nom}, "fournisseurs")
    return fournisseur_obj

@api_router.put("/fournisseurs/{fournisseur_id}", response_model=Fournisseur)
async def update_fournisseur(fournisseur_id: str, fournisseur_update: FournisseurCreate):
    result = await db.fournisseurs.update_one(
        {"id": fournisseur_id}, 
        {"$set": fournisseur_update.dict()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
    
    updated_fournisseur = await db.fournisseurs.find_one({"id": fournisseur_id})
    await log_action("admin", "update_fournisseur", {"fournisseur_id": fournisseur_id}, "fournisseurs")
    return Fournisseur(**updated_fournisseur)

@api_router.delete("/fournisseurs/{fournisseur_id}")
async def delete_fournisseur(fournisseur_id: str):
    fournisseur = await db.fournisseurs.find_one({"id": fournisseur_id})
    if not fournisseur:
        raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
    
    result = await db.fournisseurs.delete_one({"id": fournisseur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
    
    await log_action("admin", "delete_fournisseur", {"fournisseur_id": fournisseur_id, "nom": fournisseur["nom"]}, "fournisseurs")
    return {"message": "Fournisseur supprimé avec succès"}

# Enhanced stock alerts with AI predictions
@api_router.get("/stock-alerts")
async def get_stock_alerts():
    materials = await db.materials.find().to_list(1000)
    alerts = []
    
    for material in materials:
        level = "normal"
        if material["quantite"] <= material.get("seuil_critique", 5):
            level = "critique"
        elif material["quantite"] <= material.get("seuil_alerte", 10):
            level = "bas"
        
        # Get AI prediction
        prediction = await stock_predictor.predict_demand(material["id"])
        
        # Convert to Material model to ensure proper serialization
        material_obj = Material(**material)
        alert_data = {
            "material": material_obj.dict(),
            "level": level,
            "prediction": prediction
        }
        alerts.append(alert_data)
    
    return alerts

# Advanced Analytics APIs
@api_router.get("/analytics/dashboard")
async def get_analytics_dashboard():
    """Get comprehensive dashboard analytics"""
    # Get basic counts
    materials_count = await db.materials.count_documents({})
    agents_count = await db.agents.count_documents({})
    superviseurs_count = await db.superviseurs.count_documents({})
    demandes_count = await db.demandes_sortie.count_documents({})
    
    # Get stock value
    materials = await db.materials.find().to_list(1000)
    total_value = sum(mat.get("quantite", 0) * mat.get("prix_unitaire", 0) for mat in materials)
    
    # Get recent activity
    recent_demandes = await db.demandes_sortie.find().sort("date", -1).limit(10).to_list(10)
    
    # Get low stock items
    low_stock = [mat for mat in materials if mat["quantite"] <= mat.get("seuil_alerte", 10)]
    critical_stock = [mat for mat in materials if mat["quantite"] <= mat.get("seuil_critique", 5)]
    
    # Monthly trends
    now = datetime.utcnow()
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_demandes = await db.demandes_sortie.count_documents({"date": {"$gte": start_month}})
    
    # Calculate monthly value
    monthly_value = 0
    for demande in await db.demandes_sortie.find({"date": {"$gte": start_month}}).to_list(1000):
        monthly_value += demande.get("valeur_totale", 0)
    
    return {
        "summary": {
            "materials_count": materials_count,
            "agents_count": agents_count,
            "superviseurs_count": superviseurs_count,
            "demandes_count": demandes_count,
            "total_stock_value": total_value,
            "low_stock_count": len(low_stock),
            "critical_stock_count": len(critical_stock),
            "monthly_demandes": monthly_demandes,
            "monthly_value": monthly_value
        },
        "recent_activity": recent_demandes,
        "stock_alerts": {
            "low_stock": low_stock,
            "critical_stock": critical_stock
        }
    }

@api_router.get("/analytics/trends")
async def get_trends_analytics():
    """Get trends and patterns analytics"""
    # Get all demandes for analysis
    demandes = await db.demandes_sortie.find().to_list(1000)
    
    if not demandes:
        return {"error": "No data available"}
    
    # Analyze by month
    monthly_data = {}
    material_usage = {}
    
    for demande in demandes:
        month_key = demande["date"].strftime("%Y-%m")
        if month_key not in monthly_data:
            monthly_data[month_key] = {"count": 0, "value": 0}
        
        monthly_data[month_key]["count"] += 1
        monthly_data[month_key]["value"] += demande.get("valeur_totale", 0)
        
        # Track material usage
        for material_id, quantite in demande["materiels_demandes"].items():
            if material_id not in material_usage:
                material_usage[material_id] = 0
            material_usage[material_id] += quantite
    
    # Get material names for top usage
    top_materials = sorted(material_usage.items(), key=lambda x: x[1], reverse=True)[:10]
    top_materials_with_names = []
    
    for material_id, usage in top_materials:
        material = await db.materials.find_one({"id": material_id})
        if material:
            top_materials_with_names.append({
                "material_id": material_id,
                "nom": material["nom"],
                "usage": usage
            })
    
    return {
        "monthly_trends": monthly_data,
        "top_materials": top_materials_with_names,
        "total_requests": len(demandes),
        "average_value_per_request": sum(d.get("valeur_totale", 0) for d in demandes) / len(demandes) if demandes else 0
    }

# AI Predictions API
@api_router.get("/predictions/{material_id}")
async def get_stock_prediction(material_id: str, days_ahead: int = 30):
    """Get AI stock prediction for a material"""
    prediction = await stock_predictor.predict_demand(material_id, days_ahead)
    
    # Get current stock
    material = await db.materials.find_one({"id": material_id})
    if not material:
        raise HTTPException(status_code=404, detail="Matériel non trouvé")
    
    current_stock = material["quantite"]
    predicted_usage = prediction.get("prediction", 0)
    
    # Calculate reorder recommendations
    days_until_stockout = current_stock / (predicted_usage / days_ahead) if predicted_usage > 0 else float('inf')
    should_reorder = days_until_stockout < 30
    
    return {
        "material_id": material_id,
        "material_name": material["nom"],
        "current_stock": current_stock,
        "predicted_usage_" + str(days_ahead) + "_days": predicted_usage,
        "days_until_stockout": min(days_until_stockout, 365),
        "should_reorder": should_reorder,
        "confidence": prediction.get("confidence", "low"),
        "recommendation": "Commande recommandée" if should_reorder else "Stock suffisant"
    }

# Export APIs - Multi-format support
@api_router.post("/export/demandes")
async def export_demandes(export_request: ExportRequest):
    """Export demandes in multiple formats"""
    # Get data
    query = {}
    if export_request.date_debut and export_request.date_fin:
        query["date"] = {
            "$gte": export_request.date_debut,
            "$lte": export_request.date_fin
        }
    
    demandes = await db.demandes_sortie.find(query).sort("date", -1).to_list(1000)
    
    if export_request.format == "csv":
        return await export_demandes_csv(demandes)
    elif export_request.format == "excel":
        return await export_demandes_excel(demandes)
    elif export_request.format == "pdf":
        return await export_demandes_pdf(demandes, export_request.include_graphiques)
    elif export_request.format == "png":
        return await export_demandes_png(demandes)
    else:
        raise HTTPException(status_code=400, detail="Format non supporté")

async def export_demandes_csv(demandes):
    """Export to CSV format"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Date", "Superviseur", "Agent 1", "Agent 2", 
        "Matériels", "Valeur Total", "Status", "Notes"
    ])
    
    # Data
    for demande in demandes:
        materiels_str = "; ".join([f"{k}: {v}" for k, v in demande["materiels_demandes"].items()])
        writer.writerow([
            demande["date"].strftime("%Y-%m-%d %H:%M"),
            demande["superviseur_nom"],
            demande["agent1_nom"],
            demande["agent2_nom"],
            materiels_str,
            demande.get("valeur_totale", 0),
            demande["status"],
            demande.get("notes", "")
        ])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=demandes.csv"}
    )

async def export_demandes_excel(demandes):
    """Export to Excel format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demandes de Sortie"
    
    # Headers
    headers = ["Date", "Superviseur", "Agent 1", "Agent 2", "Matériels", "Valeur Total", "Status", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, demande in enumerate(demandes, 2):
        materiels_str = "; ".join([f"{k}: {v}" for k, v in demande["materiels_demandes"].items()])
        ws.cell(row=row, column=1, value=demande["date"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=2, value=demande["superviseur_nom"])
        ws.cell(row=row, column=3, value=demande["agent1_nom"])
        ws.cell(row=row, column=4, value=demande["agent2_nom"])
        ws.cell(row=row, column=5, value=materiels_str)
        ws.cell(row=row, column=6, value=demande.get("valeur_totale", 0))
        ws.cell(row=row, column=7, value=demande["status"])
        ws.cell(row=row, column=8, value=demande.get("notes", ""))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=demandes.xlsx"}
    )

async def export_demandes_pdf(demandes, include_graphiques=True):
    """Export to PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("Rapport des Demandes de Sortie", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Summary
    total_value = sum(d.get("valeur_totale", 0) for d in demandes)
    summary_text = f"""
    <b>Résumé:</b><br/>
    Nombre total de demandes: {len(demandes)}<br/>
    Valeur totale: {total_value:.2f} €<br/>
    Période: {demandes[-1]["date"].strftime("%Y-%m-%d") if demandes else "N/A"} - {demandes[0]["date"].strftime("%Y-%m-%d") if demandes else "N/A"}
    """
    summary = Paragraph(summary_text, styles['Normal'])
    story.append(summary)
    story.append(Spacer(1, 20))
    
    # Table data
    data = [["Date", "Superviseur", "Agent 1", "Valeur", "Status"]]
    for demande in demandes[:50]:  # Limit to 50 for PDF
        data.append([
            demande["date"].strftime("%Y-%m-%d"),
            demande["superviseur_nom"][:20],
            demande["agent1_nom"][:20],
            f"{demande.get('valeur_totale', 0):.2f} €",
            demande["status"]
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Generate chart if requested
    if include_graphiques and demandes:
        story.append(Spacer(1, 30))
        chart_title = Paragraph("Évolution des Demandes", styles['Heading2'])
        story.append(chart_title)
        
        # Create a simple chart using matplotlib
        fig, ax = plt.subplots(figsize=(8, 4))
        dates = [d["date"] for d in demandes[-30:]]  # Last 30 days
        values = [d.get("valeur_totale", 0) for d in demandes[-30:]]
        
        ax.plot(dates, values, marker='o')
        ax.set_title("Évolution des Valeurs de Demandes (30 derniers jours)")
        ax.set_xlabel("Date")
        ax.set_ylabel("Valeur (€)")
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Save chart to buffer
        chart_buffer = BytesIO()
        plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
        chart_buffer.seek(0)
        plt.close()
        
        # Add chart to PDF
        chart_img = Image(chart_buffer, width=6*inch, height=3*inch)
        story.append(chart_img)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=demandes.pdf"}
    )

async def export_demandes_png(demandes):
    """Export to PNG chart format"""
    if not demandes:
        raise HTTPException(status_code=400, detail="Aucune donnée à exporter")
    
    # Create comprehensive dashboard chart
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
    
    # Chart 1: Timeline of requests
    dates = [d["date"] for d in demandes[-30:]]
    values = [d.get("valeur_totale", 0) for d in demandes[-30:]]
    ax1.plot(dates, values, marker='o', color='blue')
    ax1.set_title("Évolution des Valeurs (30 derniers jours)")
    ax1.set_ylabel("Valeur (€)")
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
    
    # Chart 2: Top supervisors
    superviseurs = {}
    for d in demandes:
        sup = d["superviseur_nom"]
        superviseurs[sup] = superviseurs.get(sup, 0) + 1
    
    top_sups = sorted(superviseurs.items(), key=lambda x: x[1], reverse=True)[:10]
    if top_sups:
        names, counts = zip(*top_sups)
        ax2.bar(range(len(names)), counts, color='green')
        ax2.set_title("Top 10 Superviseurs")
        ax2.set_ylabel("Nombre de demandes")
        ax2.set_xticks(range(len(names)))
        ax2.set_xticklabels([n[:10] for n in names], rotation=45)
    
    # Chart 3: Monthly distribution
    monthly = {}
    for d in demandes:
        month = d["date"].strftime("%Y-%m")
        monthly[month] = monthly.get(month, 0) + 1
    
    if monthly:
        months, counts = zip(*sorted(monthly.items()))
        ax3.bar(months, counts, color='orange')
        ax3.set_title("Distribution Mensuelle")
        ax3.set_ylabel("Nombre de demandes")
        plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
    
    # Chart 4: Value distribution
    values_hist = [d.get("valeur_totale", 0) for d in demandes if d.get("valeur_totale", 0) > 0]
    if values_hist:
        ax4.hist(values_hist, bins=20, color='red', alpha=0.7)
        ax4.set_title("Distribution des Valeurs")
        ax4.set_xlabel("Valeur (€)")
        ax4.set_ylabel("Fréquence")
    
    plt.tight_layout()
    
    # Save to buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)
    plt.close()
    
    return StreamingResponse(
        buffer,
        media_type="image/png",
        headers={"Content-Disposition": "attachment; filename=dashboard.png"}
    )

# Historique and audit trail
@api_router.get("/historique")
async def get_historique(limit: int = 100):
    """Get action history for audit trail"""
    historique = await db.historique_actions.find().sort("date", -1).limit(limit).to_list(limit)
    return historique

# QR Code scanner route
@api_router.get("/qr/{material_id}")
async def get_material_by_qr(material_id: str):
    """Get material info by QR code scan"""
    # Extract material ID from QR data (format: MAT-{id})
    if material_id.startswith("MAT-"):
        actual_id = material_id[4:]
    else:
        actual_id = material_id
    
    material = await db.materials.find_one({"id": actual_id})
    if not material:
        raise HTTPException(status_code=404, detail="Matériel non trouvé")
    
    await log_action("user", "qr_scan", {"material_id": actual_id, "material_nom": material["nom"]}, "materials")
    return Material(**material)

# Include the router in the main app
app.include_router(api_router)

# Mount Socket.IO
app.mount("/ws", socket_app)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Initialize AI model on startup"""
    logger.info("Starting Advanced Stock Manager...")
    # Train AI model in background
    asyncio.create_task(stock_predictor.train_model())

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)